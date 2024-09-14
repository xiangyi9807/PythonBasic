from openpyxl import load_workbook

wb = load_workbook("excel_test_3.xlsx")

# 打印所有sheet
print(wb.sheetnames)

# 切换sheet
sheet = wb["Sheet1"]

# 打印单元格的值
print(sheet["C3"].value)
print("=" * 50)


# 获取多个单元格的值
sheet = wb["Sheet"]
for cell in sheet["B2:B10"]:
    print(cell[0].value)
print("=" * 50)


# 按行遍历获取全部数据
for row in sheet:
    for cell in row:
        print(cell.value, end="\t")
    print()
print("=" * 50)


# 按列遍历获取数据
for column in sheet.columns:
    for cell in column:
        print(cell.value, end="\t")
    print()
print("=" * 50)


# 获取指定行指定列的数据
for row in sheet.iter_rows(min_row=2, max_row=10, min_col=2, max_col=3):
    for cell in row:
        print(cell.value, end="\t")
    print()
print("=" * 50)




