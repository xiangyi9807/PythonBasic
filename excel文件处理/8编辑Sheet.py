#! /usr/bin/env python
# author: vin

from openpyxl import Workbook

wb=Workbook()

sheet = wb.active   # 获取当前sheet

sheet = wb.create_sheet("Sheet1")   # 新建sheet

sheet = wb["Sheet1"]    # 切换sheet

sheet.title = "Test"    # 修改sheet名

print(sheet.title)  # 打印sheet名

sheet = wb["Sheet"]     # 获取sheet

wb.move_sheet('Sheet', 1)   # 移动sheet，正数表示向右移动，复数表示向左移动，数字表示移动量

print(wb.sheetnames)    # 打印所有sheet

cp_sheet=wb.copy_worksheet("Sheet")     # 复制sheet

del wb["cp_sheet"]

wb.save("excel_test_6.xlsx")

