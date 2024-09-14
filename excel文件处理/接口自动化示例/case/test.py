#! /usr/bin/env python
# author: vin

import requests
from urllib import parse
import json
import os
import openpyxl

data_path = os.path.dirname(os.path.dirname(os.path.abspath('__file__'))) + '\\data\\accounts.xlsx'

try:
    wb = openpyxl.load_workbook(data_path)
except Exception as e:
    raise e

sheet = wb.active
account_dict = {}
id_list = []
name_list = []

for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=2):
    i = 1
    for cell in row:
        if i%2 != 0:
            id_list.append(cell.value)
        else:
            name_list.append(cell.value)
        i += 1

account_dict = zip(id_list, name_list)

result_path = data_path = os.path.dirname(os.path.dirname(os.path.abspath('__file__'))) + '\\result\\result.xlsx'

wb = openpyxl.load_workbook(result_path)

sheet = wb.active

i = 1
for key,value in account_dict:
    userid = key
    name = value

    url = "https://dev.oc.cnpc.com.cn:2001/core/cnpcoa-auth-center/custom/get_token"
    body = '{"userid":"%s", "name":"%s", "device":"meeting"}' % (userid, name)
    payload = 'user_info=' + parse.quote(body)
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Cookie': 'sl-session=RaUhdAp4hmX09LJjlCdRtA=='
    }

    response = requests.request("POST", url, headers=headers, data=payload, verify=False)
    result = json.loads(response.text)

    cell = 'A%d' % i
    sheet[cell] = result['data']['token']
    i += 1

wb.save(result_path)






